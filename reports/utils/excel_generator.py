# reports/utils/excel_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _


def generate_excel_report(tickets, filters, report_type, stats_data=None):
    """Excel hisobot yaratish"""
    
    # Workbook yaratish
    wb = Workbook()
    
    # Default sheet ni o'chirish
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Report type bo'yicha
    if report_type == 'tickets':
        create_tickets_sheet(wb, tickets, filters)
    
    elif report_type == 'statistics':
        create_statistics_sheets(wb, stats_data, filters)
    
    elif report_type == 'technician_performance':
        create_technician_performance_sheet(wb, stats_data, filters)
    
    elif report_type == 'system_analysis':
        create_system_analysis_sheet(wb, stats_data, filters)
    
    elif report_type == 'regional_analysis':
        create_regional_analysis_sheet(wb, stats_data, filters)
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def create_tickets_sheet(wb, tickets, filters):
    """Ticketlar ro'yxati sheet"""
    
    ws = wb.create_sheet(_("Murojaatlar"))
    
    # Header
    ws.append([_("IIV TEXNIK MUROJAATLAR HISOBOTI")])
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Sana
    ws.append([f"{_('Yaratildi')}: {timezone.now().strftime('%d.%m.%Y %H:%M')}"])
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=10, color='64748b')
    
    # Filter info
    filter_text = get_filter_text(filters)
    if filter_text:
        ws.append([f"{_('Filtrlar')}: {filter_text}"])
        ws.merge_cells('A3:H3')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A3'].font = Font(size=10, italic=True)
    
    # Bo'sh qator
    ws.append([])
    
    # Table headers
    headers = [
        _('ID'),
        _('Sana'),
        _('Tizim'),
        _('Viloyat'),
        _('Foydalanuvchi'),
        _('Holat'),
        _('Mas\'ul xodim'),
        _('Baho'),
    ]
    
    header_row = ws.max_row + 1
    ws.append(headers)
    
    # Header style
    header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
    
    # Data rows
    for ticket in tickets:
        ws.append([
            ticket.get_ticket_number(),
            ticket.created_at.strftime('%d.%m.%Y %H:%M'),
            ticket.system.name,
            ticket.region.name if ticket.region else '-',
            ticket.user.get_full_name(),
            ticket.get_status_display(),
            ticket.assigned_to.get_full_name() if ticket.assigned_to else '-',
            f"{ticket.rating}⭐" if ticket.rating else '-',
        ])
        
        # Row style
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = get_border()
            cell.alignment = Alignment(vertical='center')
            
            # Zebra stripes
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    
    # Column widths
    column_widths = [15, 18, 25, 20, 30, 20, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Summary
    ws.append([])
    ws.append([f"{_('Jami')}: {tickets.count()} {_('ta murojaat')}"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
    ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='center')


def create_statistics_sheets(wb, stats_data, filters):
    """Statistika sheets (bir nechta sheet)"""
    
    # Status bo'yicha
    if stats_data.get('by_status'):
        ws = wb.create_sheet(_("Holat bo'yicha"))
        add_simple_stat_table(
            ws, 
            _("Holat bo'yicha statistika"),
            [_('Holat'), _('Soni')],
            [(item['status'], item['count']) for item in stats_data['by_status']]
        )
    
    # Tizim bo'yicha
    if stats_data.get('by_system'):
        ws = wb.create_sheet(_("Tizimlar"))
        add_simple_stat_table(
            ws,
            _("Tizimlar bo'yicha statistika"),
            [_('Tizim'), _('Soni')],
            [(item['system__name'], item['count']) for item in stats_data['by_system']]
        )
    
    # Viloyat bo'yicha
    if stats_data.get('by_region'):
        ws = wb.create_sheet(_("Viloyatlar"))
        add_simple_stat_table(
            ws,
            _("Viloyatlar bo'yicha statistika"),
            [_('Viloyat'), _('Soni')],
            [(item['region__name'] or '-', item['count']) for item in stats_data['by_region']]
        )
    
    # Ustuvorlik bo'yicha
    if stats_data.get('by_priority'):
        ws = wb.create_sheet(_("Ustuvorlik"))
        add_simple_stat_table(
            ws,
            _("Ustuvorlik bo'yicha statistika"),
            [_('Ustuvorlik'), _('Soni')],
            [(item['priority'], item['count']) for item in stats_data['by_priority']]
        )
    
    # Baholash bo'yicha
    if stats_data.get('by_rating'):
        ws = wb.create_sheet(_("Baholash"))
        add_simple_stat_table(
            ws,
            _("Baholash bo'yicha statistika"),
            [_('Baho'), _('Soni')],
            [(f"{item['rating']}⭐", item['count']) for item in stats_data['by_rating']]
        )


def create_technician_performance_sheet(wb, performance_data, filters):
    """Texniklar samaradorligi sheet"""
    
    ws = wb.create_sheet(_("Texniklar samaradorligi"))
    
    # Title
    ws.append([_("TEXNIKLAR SAMARADORLIGI")])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.append([])
    
    # Headers
    headers = [
        _('Texnik'),
        _('Jami biriktirilgan'),
        _('Hal qilingan'),
        _('Jarayonda'),
        _('Qayta ochilgan'),
        _('O\'rtacha baho'),
    ]
    
    header_row = ws.max_row + 1
    ws.append(headers)
    
    # Header style
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
    
    # Data
    for item in performance_data:
        full_name = f"{item['assigned_to__first_name']} {item['assigned_to__last_name']}"
        ws.append([
            full_name,
            item['total_assigned'],
            item['resolved'],
            item['in_progress'],
            item['reopened'],
            round(item['avg_rating'], 2) if item['avg_rating'] else '-',
        ])
        
        # Style
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    
    # Column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25


def create_system_analysis_sheet(wb, analysis_data, filters):
    """Tizimlar tahlili sheet"""
    
    ws = wb.create_sheet(_("Tizimlar tahlili"))
    
    # Title
    ws.append([_("TIZIMLAR TAHLILI")])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.append([])
    
    # Headers
    headers = [
        _('Tizim'),
        _('Jami'),
        _('Hal qilingan'),
        _('O\'rtacha baho'),
        _('Yuqori ustuvorlik'),
    ]
    
    header_row = ws.max_row + 1
    ws.append(headers)
    
    # Header style
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
    
    # Data
    for item in analysis_data:
        ws.append([
            item['system__name'],
            item['total'],
            item['resolved'],
            round(item['avg_rating'], 2) if item['avg_rating'] else '-',
            item['high_priority'],
        ])
        
        # Style
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def create_regional_analysis_sheet(wb, analysis_data, filters):
    """Viloyatlar tahlili sheet"""
    
    ws = wb.create_sheet(_("Viloyatlar tahlili"))
    
    # Title
    ws.append([_("VILOYATLAR TAHLILI")])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.append([])
    
    # Headers
    headers = [
        _('Viloyat'),
        _('Jami'),
        _('Hal qilingan'),
        _('Jarayonda'),
        _('O\'rtacha baho'),
    ]
    
    header_row = ws.max_row + 1
    ws.append(headers)
    
    # Header style
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
    
    # Data
    for item in analysis_data:
        ws.append([
            item['region__name'] or '-',
            item['total'],
            item['resolved'],
            item['in_progress'],
            round(item['avg_rating'], 2) if item['avg_rating'] else '-',
        ])
        
        # Style
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


# ============================================
# HELPER FUNCTIONS
# ============================================

def add_simple_stat_table(ws, title, headers, data):
    """Oddiy statistika jadvali"""
    
    # Title
    ws.append([title])
    ws.merge_cells(f'A1:B1')
    ws['A1'].font = Font(size=14, bold=True, color='1e40af')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.append([])
    
    # Headers
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
    
    # Data
    for row_data in data:
        ws.append(row_data)
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_num % 2 == 1:
                cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15


def get_border():
    """Border style"""
    return Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db')
    )


def get_filter_text(filters):
    """Filter ma'lumotlari text"""
    parts = []
    
    if filters.get('date_from') and filters.get('date_to'):
        parts.append(f"{filters['date_from'].strftime('%d.%m.%Y')} - {filters['date_to'].strftime('%d.%m.%Y')}")
    
    if filters.get('system'):
        parts.append(f"{_('Tizim')}: {filters['system'].name}")
    
    if filters.get('region'):
        parts.append(f"{_('Viloyat')}: {filters['region'].name}")
    
    if filters.get('status'):
        parts.append(f"{_('Holat')}: {filters['status']}")
    
    if filters.get('assigned_to'):
        parts.append(f"{_('Mas\'ul')}: {filters['assigned_to'].get_full_name()}")
    
    return " • ".join(parts)